from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import csv
import io
import uuid
from openpyxl import load_workbook

# queue + storage + progress
from app.queue.enqueue import enqueue_urls
from app.storage.sqlite import fetch_results
from app.core.progress import set_run_progress, get_run_progress


router = APIRouter()


# -------------------------------
# Helpers: CSV & Excel parsing
# -------------------------------

def extract_csv(raw: bytes):
    for enc in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue

    reader = csv.reader(io.StringIO(text))
    urls = []

    for row in reader:
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                urls.append(cell.strip())
                break

    return urls


def extract_excel(raw: bytes):
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = wb.active

        urls = []

        for row in sheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str):
                url = row[0].strip()
                if url:
                    urls.append(url)

        wb.close()
        return urls

    except Exception:
        raise HTTPException(400, "Invalid Excel file format")


# -------------------------------
# MAIN ENDPOINTS
# -------------------------------

@router.post("/api/analyze")
async def analyze(file: UploadFile = File(...)):
    raw = await file.read()
    filename = file.filename.lower()

    # handle CSV vs Excel
    if filename.endswith(".csv"):
        urls = extract_csv(raw)
    else:
        urls = extract_excel(raw)

    if not urls:
        raise HTTPException(status_code=400, detail="No valid URLs found")

    # create run id
    run_id = str(uuid.uuid4())

  # init progress BEFORE enqueue!
    # BEFORE enqueue
    total = enqueue_urls(run_id, urls)

    if total == 0:
        return {"error": "No valid URLs"}

    # Initialize progress BEFORE worker starts
    set_run_progress(run_id, processed=0, total=total)

    return {"run_id": run_id, "total": total}


    


@router.get("/api/progress/{run_id}")
def progress(run_id: str):
    return get_run_progress(run_id)


@router.get("/api/results/{run_id}")
def results(run_id: str):
    rows = fetch_results(run_id)

    formatted = []

    for r in rows:
        # r index reference:
        # 0 run_id
        # 1 url
        # 2 niche
        # 3 cms
        # 4 submission
        # 5 authority
        # 6 spam
        # 7 category
        # 8 structure
        # 9 tier

        formatted.append({
            "url": r[1],
            "niche": r[2],
            "cms": r[3],
            "submission_type": r[4],
            "guest_post": True if r[4] in ("direct", "email", "registration") else False,
            "spam_risk": "High" if r[6] >= 0.8 else "Medium" if r[6] >= 0.5 else "Low",
            "authority_score": r[5],
            "quality_tier": r[9],
            "structure_group": r[8],
            "index_status": True,              # placeholder
            "content_length": 0,               # can be added later
            "load_time": 0,                    # optional future add
        })

    return JSONResponse(content=formatted)
